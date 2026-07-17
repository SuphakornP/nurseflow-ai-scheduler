from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.export import build_workbook
from app.models import (
    DailyRequest,
    Nurse,
    OptimizationProfile,
    RequestConstraintMode,
    RequestResolution,
    ScheduleProblem,
    ScheduleRules,
    ShiftCode,
    SkillLevel,
)
from app.solver import generate_schedule
from app.validation import validate_schedule

from conftest import make_assignments


@pytest.fixture
def request_nurses() -> list[Nurse]:
    return [
        Nurse(id=identifier, nickname=f"Nurse {identifier}", skill_level=level)
        for identifier, level in (
            ("A", SkillLevel.INCHARGE),
            ("B", SkillLevel.INCHARGE),
            ("C", SkillLevel.INCHARGE),
            ("D", SkillLevel.INCHARGE),
            ("E", SkillLevel.TRAINEE_INC),
            ("F", SkillLevel.TRAINEE_INC),
            ("G", SkillLevel.TRAINEE_INC),
        )
    ]


def _problem_with_requests(
    request_nurses,
    permissive_skill_ranges,
    requests: list[DailyRequest],
) -> ScheduleProblem:
    return ScheduleProblem(
        period_name="Request semantics fixture",
        period_start=date(2026, 8, 1),
        period_end=date(2026, 8, 1),
        nurses=request_nurses,
        requests=requests,
        rules=ScheduleRules(
            day_staffing=3,
            night_staffing=3,
            skill_ranges=permissive_skill_ranges,
        ),
        time_limit_seconds=2,
    )


def _request(
    nurse_id: str,
    raw_value: str,
    *,
    mode: RequestConstraintMode = RequestConstraintMode.PREFERENCE,
    resolution: RequestResolution | None = None,
) -> DailyRequest:
    return DailyRequest(
        nurse_id=nurse_id,
        request_date=date(2026, 8, 1),
        raw_value=raw_value,
        constraint_mode=mode,
        resolution=resolution,
    )


def test_conflicting_shift_requests_remain_feasible_and_are_reported(
    request_nurses,
    permissive_skill_ranges,
) -> None:
    problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [_request(identifier, "D") for identifier in ("A", "B", "C", "D")],
    )

    result = generate_schedule(problem)

    assert result.status in {"OPTIMAL", "FEASIBLE"}
    assert result.validation is not None and result.validation.is_valid
    assert result.metrics["REQUEST_COUNT"] == 4
    assert result.metrics["REQUEST_SATISFIED_COUNT"] == 3
    assert result.metrics["REQUEST_SATISFACTION_RATE"] == 0.75


def test_requests_first_freezes_overall_preference_satisfaction(
    request_nurses,
    permissive_skill_ranges,
) -> None:
    problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [_request(identifier, "D") for identifier in ("A", "B", "C", "D")],
    )
    problem.optimization_profile = OptimizationProfile.REQUESTS_FIRST

    result = generate_schedule(problem)

    assert any(
        phase.name == "MAXIMIZE_REQUEST_PREFERENCES" for phase in result.phases
    )
    assert result.metrics["REQUEST_SATISFIED_COUNT"] == 3


@pytest.mark.parametrize("requested", [ShiftCode.VACATION, ShiftCode.EDUCATION])
def test_vacation_and_education_preferences_are_selectable_without_being_locks(
    requested: ShiftCode,
    request_nurses,
    permissive_skill_ranges,
) -> None:
    problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [_request("A", requested.value)],
    )

    result = generate_schedule(problem)
    assigned = next(item for item in result.assignments if item.nurse_id == "A")

    assert result.status in {"OPTIMAL", "FEASIBLE"}
    assert assigned.shift == requested
    assert result.metrics["REQUEST_SATISFIED_COUNT"] == 1


def test_human_resolved_allowed_set_is_a_soft_preference(
    request_nurses,
    permissive_skill_ranges,
) -> None:
    problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [
            _request(
                "A",
                "D1",
                resolution=RequestResolution(
                    allowed_assignments=[ShiftCode.DAY],
                ),
            ),
            _request("B", "D"),
            _request("C", "D"),
            _request("D", "D"),
        ],
    )

    result = generate_schedule(problem)

    assert result.status in {"OPTIMAL", "FEASIBLE"}
    assert result.validation is not None and result.validation.is_valid
    assert result.metrics["REQUEST_COUNT"] == 4
    assert result.metrics["REQUEST_SATISFIED_COUNT"] == 3


def test_conflicting_admin_locks_remain_infeasible(
    request_nurses,
    permissive_skill_ranges,
) -> None:
    problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [
            _request("A", "D", mode=RequestConstraintMode.LOCKED),
            _request("B", "D", mode=RequestConstraintMode.LOCKED),
            _request("C", "D", mode=RequestConstraintMode.LOCKED),
            _request("D", "D", mode=RequestConstraintMode.LOCKED),
        ],
    )

    result = generate_schedule(problem)

    assert result.status == "INFEASIBLE"
    assert result.assignments == []


def test_validator_ignores_unmet_preference_but_rejects_unmet_lock(
    request_nurses,
    permissive_skill_ranges,
) -> None:
    preference_problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [_request("A", "D")],
    )
    assignments = make_assignments(
        preference_problem,
        {
            "A": [ShiftCode.NIGHT],
            "B": [ShiftCode.DAY],
            "C": [ShiftCode.DAY],
            "D": [ShiftCode.DAY],
            "E": [ShiftCode.NIGHT],
            "F": [ShiftCode.NIGHT],
            "G": [ShiftCode.OFF],
        },
    )

    preference_report = validate_schedule(preference_problem, assignments)
    locked_problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [_request("A", "D", mode=RequestConstraintMode.LOCKED)],
    )
    locked_report = validate_schedule(locked_problem, assignments)

    preference_check = next(
        item
        for item in preference_report.checks
        if item.code == "FIXED_CLINICAL_PRESERVED"
    )
    locked_check = next(
        item
        for item in locked_report.checks
        if item.code == "FIXED_CLINICAL_PRESERVED"
    )
    assert preference_report.is_valid
    assert preference_check.status == "PASS"
    assert not locked_report.is_valid
    assert locked_check.status == "FAIL"


def test_export_lists_unfulfilled_shift_preferences(
    request_nurses,
    permissive_skill_ranges,
) -> None:
    problem = _problem_with_requests(
        request_nurses,
        permissive_skill_ranges,
        [_request(identifier, "D") for identifier in ("A", "B", "C", "D")],
    )
    result = generate_schedule(problem)
    assert result.validation is not None and result.validation.is_valid

    content = build_workbook(
        problem,
        result.assignments,
        result.summaries,
        result.validation,
    )
    workbook = load_workbook(BytesIO(content), read_only=True)
    sheet = workbook["Unfulfilled Requests"]

    assert sheet.max_row == 2
    assert sheet["D2"].value == "D"
    assert sheet["E2"].value is None
    assert sheet["F2"].value != "D"
