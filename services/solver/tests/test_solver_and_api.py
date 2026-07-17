import base64
import logging
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.export import WorkbookTemplateError, build_workbook
from app.main import app
from app.models import OptimizationProfile, ShiftCode, SourceWorkbookTemplate


SOLVER_API_TOKEN = "solver-test-token-with-at-least-32-characters"
AUTH_HEADERS = {"Authorization": f"Bearer {SOLVER_API_TOKEN}"}


def _source_template(problem, worksheet_name: str = "คำขอเวร") -> SourceWorkbookTemplate:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_name
    worksheet["A1"] = "MICU August request template"
    worksheet["A2"] = "รหัสพนักงาน"
    worksheet["B2"] = "ชื่อ - สกุล"
    worksheet["C2"] = "Level"
    for offset, day in enumerate((29, 30, 31), start=4):
        worksheet.cell(row=2, column=offset, value=day)
    dates = [
        problem.period_start.fromordinal(problem.period_start.toordinal() + offset)
        for offset in range((problem.period_end - problem.period_start).days + 1)
    ]
    date_columns = {}
    for column, assignment_date in enumerate(dates, start=7):
        worksheet.cell(row=2, column=column, value=assignment_date.day)
        date_columns[assignment_date] = column
    worksheet.cell(row=2, column=38, value="หมายเหตุ")

    nurse_rows = {}
    for row, nurse in enumerate(problem.nurses, start=3):
        nurse_rows[nurse.id] = row
        worksheet.cell(row=row, column=1, value=None)
        worksheet.cell(row=row, column=2, value=nurse.nickname)
        worksheet.cell(row=row, column=3, value=nurse.skill_level.value)
        worksheet.cell(row=row, column=4, value="D")
        worksheet.cell(row=row, column=5, value="N")
        worksheet.cell(row=row, column=6, value="OFF")
        worksheet.cell(row=row, column=38, value=None)
    worksheet.cell(row=3, column=7).fill = PatternFill("solid", fgColor="C6E0B4")

    output = BytesIO()
    workbook.save(output)
    return SourceWorkbookTemplate(
        content_base64=base64.b64encode(output.getvalue()).decode("ascii"),
        worksheet_name=worksheet.title,
        nurse_rows=nurse_rows,
        date_columns=date_columns,
    )


def test_health_is_public_and_minimal(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("SOLVER_API_TOKEN", raising=False)
    client = TestClient(app)
    health = client.get("/health")
    assert health.status_code == 200
    assert health.json() == {"status": "ok"}


@pytest.mark.parametrize("path", ["/docs", "/redoc", "/openapi.json"])
def test_solver_schema_and_docs_are_not_public(path: str) -> None:
    response = TestClient(app).get(path)

    assert response.status_code == 404


@pytest.mark.parametrize("configured_token", [None, "too-short"])
def test_protected_routes_fail_closed_when_token_is_not_configured(
    configured_token: str | None,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    if configured_token is None:
        monkeypatch.delenv("SOLVER_API_TOKEN", raising=False)
    else:
        monkeypatch.setenv("SOLVER_API_TOKEN", configured_token)

    response = TestClient(app).get("/demo")

    assert response.status_code == 503
    assert response.json() == {"detail": "Solver authentication is not configured"}


@pytest.mark.parametrize(
    ("method", "path"),
    [("GET", "/demo"), ("POST", "/generate"), ("POST", "/export")],
)
def test_protected_routes_require_bearer_token(
    method: str,
    path: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("SOLVER_API_TOKEN", SOLVER_API_TOKEN)

    response = TestClient(app).request(method, path, json={} if method == "POST" else None)

    assert response.status_code == 401
    assert response.headers["www-authenticate"] == "Bearer"


def test_wrong_bearer_token_is_rejected(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SOLVER_API_TOKEN", SOLVER_API_TOKEN)

    response = TestClient(app).get(
        "/demo",
        headers={"Authorization": "Bearer wrong-token-that-is-also-long-enough"},
    )

    assert response.status_code == 401
    assert response.headers["www-authenticate"] == "Bearer"


def test_correct_bearer_token_allows_demo(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("SOLVER_API_TOKEN", SOLVER_API_TOKEN)
    client = TestClient(app)

    demo = client.get("/demo", headers=AUTH_HEADERS)
    assert demo.status_code == 200
    body = demo.json()
    assert len(body["nurses"]) == 28
    assert all("nickname" in nurse for nurse in body["nurses"])
    assert all("first_name" not in nurse and "last_name" not in nurse for nurse in body["nurses"])


def test_generate_input_rejection_is_sanitized_and_logged(
    caplog: pytest.LogCaptureFixture,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("SOLVER_API_TOKEN", SOLVER_API_TOKEN)
    client = TestClient(app)
    problem = client.get("/demo", headers=AUTH_HEADERS).json()
    private_value = "EMP-SECRET-77 NursePrivate raw-token-Z"
    private_nurse_id = problem["requests"][0]["nurse_id"]
    problem["requests"][0]["raw_value"] = private_value
    problem["requests"][0].pop("resolution", None)

    with caplog.at_level(logging.WARNING, logger="nurseflow.solver"):
        response = client.post(
            "/generate",
            json=problem,
            headers={**AUTH_HEADERS, "x-request-id": "safe-request-123"},
        )

    assert response.status_code == 422
    assert response.json() == {
        "detail": {
            "issue_count": 1,
            "reason_codes": ["UNRESOLVED_REQUEST"],
        }
    }
    assert "solver.generate.input_rejected" in caplog.text
    assert "safe-request-123" in caplog.text
    assert private_value not in response.text
    assert private_value not in caplog.text
    assert private_nurse_id not in response.text
    assert private_nurse_id not in caplog.text


def test_request_schema_rejection_is_sanitized_and_logged(
    caplog: pytest.LogCaptureFixture,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("SOLVER_API_TOKEN", SOLVER_API_TOKEN)
    private_value = "EMP-SECRET-88 NursePrivate schema-input"

    with caplog.at_level(logging.WARNING, logger="nurseflow.solver"):
        response = TestClient(app).post(
            "/generate",
            json={"period_start": "not-a-date", "nurses": [{"id": private_value}]},
            headers={**AUTH_HEADERS, "x-request-id": "safe-schema-123"},
        )

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert detail["reason_codes"] == ["PAYLOAD_VALIDATION"]
    assert detail["issue_count"] >= 1
    assert "solver.request.schema_rejected" in caplog.text
    assert "safe-schema-123" in caplog.text
    assert private_value not in response.text
    assert private_value not in caplog.text


def test_demo_solver_passes_independent_validation(generated_demo) -> None:
    problem, result = generated_demo
    expected_assignments = len(problem.nurses) * 31
    assert result.status in {"OPTIMAL", "FEASIBLE"}
    assert len(result.assignments) == expected_assignments
    assert result.validation is not None
    assert result.validation.is_valid
    assert [item.name for item in result.phases[:4]] == [
        "MAXIMIZE_O1",
        "MAXIMIZE_O2",
        "MAXIMIZE_O3",
        "MAXIMIZE_O4",
    ]


def test_showcase_profiles_produce_three_valid_candidates(generated_profiles) -> None:
    signatures = set()
    for profile in OptimizationProfile:
        _, result = generated_profiles[profile]
        assert result.status in {"OPTIMAL", "FEASIBLE"}
        assert result.validation is not None and result.validation.is_valid
        assert result.metrics["OPTIMIZATION_PROFILE"] == profile.value
        signatures.add(
            tuple(
                (item.nurse_id, item.assignment_date, item.shift)
                for item in result.assignments
            )
        )
    assert len(signatures) == 3


def test_excel_contains_required_audit_sheets(generated_demo) -> None:
    problem, result = generated_demo
    assert result.validation is not None
    content = build_workbook(
        problem,
        result.assignments,
        result.summaries,
        result.validation,
    )
    workbook = load_workbook(BytesIO(content), read_only=True)
    assert workbook.sheetnames == [
        "Schedule",
        "Summary",
        "Validation",
        "Unfulfilled Requests",
        "Metadata",
    ]
    assert workbook["Metadata"]["B7"].value == "Nickname and pseudonymous internal ID only"


def test_excel_preserves_sanitized_source_layout_and_styles(generated_demo) -> None:
    problem, result = generated_demo
    assert result.validation is not None
    template = _source_template(problem)

    content = build_workbook(
        problem,
        result.assignments,
        result.summaries,
        result.validation,
        template,
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    worksheet = workbook["คำขอเวร"]
    first_assignment = next(
        item
        for item in result.assignments
        if item.nurse_id == problem.nurses[0].id
        and item.assignment_date == problem.period_start
    )
    expected_label = {
        ShiftCode.DAY: "D",
        ShiftCode.NIGHT: "N",
        ShiftCode.OFF: "OFF",
        ShiftCode.VACATION: "Vac",
        ShiftCode.EDUCATION: "ชED",
    }[first_assignment.shift]

    assert workbook.sheetnames == [
        "คำขอเวร",
        "Summary",
        "Validation",
        "Unfulfilled Requests",
        "Metadata",
    ]
    assert worksheet["A1"].value == "MICU August request template"
    assert worksheet["D3"].value == "D"
    assert worksheet["G3"].value == expected_label
    assert worksheet["G3"].fill.fgColor.rgb.endswith("C6E0B4")
    assert worksheet["A3"].value is None
    assert worksheet["AL3"].value is None


def test_excel_rejects_case_insensitive_reserved_source_sheet_name(generated_demo) -> None:
    problem, result = generated_demo
    assert result.validation is not None

    with pytest.raises(WorkbookTemplateError):
        build_workbook(
            problem,
            result.assignments,
            result.summaries,
            result.validation,
            _source_template(problem, "summary"),
        )


def test_excel_escapes_formula_like_attacker_controlled_text(generated_demo) -> None:
    problem, result = generated_demo
    assert result.validation is not None

    unsafe_problem = problem.model_copy(deep=True)
    unsafe_problem.period_name = "=HYPERLINK(\"https://attacker.example\")"
    unsafe_problem.nurses[0].nickname = "  +SUM(1,1)"
    unsafe_summaries = [item.model_copy(deep=True) for item in result.summaries]
    unsafe_summaries[0].nickname = "@SUM(1,1)"
    unsafe_validation = result.validation.model_copy(deep=True)
    unsafe_validation.checks[0].details = ["\t-CMD|' /C calc'!A0"]

    content = build_workbook(
        unsafe_problem,
        result.assignments,
        unsafe_summaries,
        unsafe_validation,
    )
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)

    assert workbook["Schedule"]["A1"].value == "'=HYPERLINK(\"https://attacker.example\")"
    assert workbook["Schedule"]["B3"].value == "'  +SUM(1,1)"
    assert workbook["Summary"]["B2"].value == "'@SUM(1,1)"
    assert workbook["Validation"]["E2"].value == "'\t-CMD|' /C calc'!A0"
    assert workbook["Metadata"]["B1"].value == "'=HYPERLINK(\"https://attacker.example\")"


def test_export_endpoint_exposes_filename_header(
    generated_demo,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("SOLVER_API_TOKEN", SOLVER_API_TOKEN)
    problem, result = generated_demo
    client = TestClient(app)
    response = client.post(
        "/export",
        json={
            "problem": problem.model_dump(mode="json"),
            "assignments": [
                item.model_dump(mode="json") for item in result.assignments
            ],
        },
        headers=AUTH_HEADERS,
    )
    assert response.status_code == 200
    assert response.headers["x-filename"].endswith(".xlsx")
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_export_endpoint_sanitizes_filename_headers(
    generated_demo,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("SOLVER_API_TOKEN", SOLVER_API_TOKEN)
    problem, result = generated_demo
    payload = problem.model_dump(mode="json")
    payload["period_name"] = 'Quarter "\r\nX-Evil: yes/../../'

    response = TestClient(app).post(
        "/export",
        json={
            "problem": payload,
            "assignments": [item.model_dump(mode="json") for item in result.assignments],
        },
        headers=AUTH_HEADERS,
    )

    assert response.status_code == 200
    assert response.headers["x-filename"] == "nurseflow-quarter-x-evil-yes.xlsx"
    assert response.headers["content-disposition"] == (
        'attachment; filename="nurseflow-quarter-x-evil-yes.xlsx"'
    )
