from __future__ import annotations

import base64
import binascii
from io import BytesIO
from zipfile import BadZipFile, ZIP_DEFLATED, ZIP_STORED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import (
    Assignment,
    NurseSummary,
    RequestConstraintMode,
    ScheduleProblem,
    ShiftCode,
    SourceWorkbookTemplate,
    ValidationReport,
)
from .normalization import normalize_request


NAVY = "17324D"
TEAL = "00A6A6"
PALE_TEAL = "DDF5F2"
PALE_BLUE = "E8F1F8"
PALE_YELLOW = "FFF4CC"
PALE_RED = "FDE8E7"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="D7E0E8")
TEMPLATE_SHIFT_LABELS = {
    ShiftCode.DAY: "D",
    ShiftCode.NIGHT: "N",
    ShiftCode.OFF: "OFF",
    ShiftCode.VACATION: "Vac",
    ShiftCode.EDUCATION: "ชED",
}
AUDIT_SHEET_NAMES = ("Summary", "Validation", "Unfulfilled Requests", "Metadata")
AUDIT_SHEET_NAMES_CASEFOLD = {name.casefold() for name in AUDIT_SHEET_NAMES}
MAX_TEMPLATE_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_TEMPLATE_ZIP_ENTRIES = 512
MAX_TEMPLATE_COMPRESSION_RATIO = 300


class WorkbookTemplateError(ValueError):
    pass


def _forbidden_archive_part(name: str) -> bool:
    normalized = name.casefold()
    return (
        normalized == "xl/connections.xml"
        or normalized.endswith("/vbaproject.bin")
        or normalized.startswith("xl/externallinks/")
        or normalized.startswith("xl/embeddings/")
        or normalized.startswith("xl/activex/")
        or normalized.startswith("xl/ctrlprops/")
        or normalized.startswith("xl/querytables/")
        or normalized.startswith("xl/pivotcache/")
        or normalized.startswith("xl/media/")
        or normalized.startswith("customxml/")
        or normalized.startswith("_xmlsignatures/")
        or (
            normalized.startswith("xl/drawings/drawing")
            and normalized.endswith(".xml")
        )
    )


def _assert_safe_template_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
    except (BadZipFile, OSError, ValueError) as exc:
        raise WorkbookTemplateError("The sanitized source workbook is not a valid ZIP") from exc

    if not 1 <= len(entries) <= MAX_TEMPLATE_ZIP_ENTRIES:
        raise WorkbookTemplateError("The source workbook contains too many ZIP entries")
    names: set[str] = set()
    total_uncompressed = 0
    for entry in entries:
        name = entry.filename
        segments = name.split("/")
        if (
            not name
            or name.startswith("/")
            or "\\" in name
            or "\x00" in name
            or ".." in segments
            or name in names
        ):
            raise WorkbookTemplateError("The source workbook contains an unsafe ZIP path")
        if entry.flag_bits & 0x1 or entry.compress_type not in {
            ZIP_STORED,
            ZIP_DEFLATED,
        }:
            raise WorkbookTemplateError("Encrypted or unsupported ZIP entries are not allowed")
        if _forbidden_archive_part(name):
            raise WorkbookTemplateError("Active or embedded workbook content is not allowed")
        names.add(name)
        total_uncompressed += entry.file_size
        if total_uncompressed > MAX_TEMPLATE_UNCOMPRESSED_BYTES:
            raise WorkbookTemplateError("The expanded source workbook exceeds 64 MB")
        if (
            entry.file_size > 1024 * 1024
            and (
                entry.compress_size == 0
                or entry.file_size / entry.compress_size
                > MAX_TEMPLATE_COMPRESSION_RATIO
            )
        ):
            raise WorkbookTemplateError("The source workbook compression ratio is unsafe")
    if not {
        "[Content_Types].xml",
        "xl/workbook.xml",
    } <= names or not any(
        name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        for name in names
    ):
        raise WorkbookTemplateError("The source archive is not an xlsx workbook")


def _safe_text(value: str) -> str:
    # Excel-compatible readers may treat formula markers after leading whitespace
    # or control characters as executable content.
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _header(cell, fill: str = NAVY) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)


def _body(cell, fill: str | None = None, centered: bool = True) -> None:
    if isinstance(cell.value, str):
        cell.value = _safe_text(cell.value)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal="center" if centered else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)


def _period_dates(problem: ScheduleProblem):
    return [
        problem.period_start.fromordinal(problem.period_start.toordinal() + offset)
        for offset in range((problem.period_end - problem.period_start).days + 1)
    ]


def _load_source_schedule(
    problem: ScheduleProblem,
    assignments: list[Assignment],
    template: SourceWorkbookTemplate,
) -> tuple[Workbook, Worksheet]:
    try:
        content = base64.b64decode(template.content_base64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise WorkbookTemplateError("The source workbook encoding is invalid") from exc
    _assert_safe_template_archive(content)
    try:
        workbook = load_workbook(BytesIO(content), keep_links=False)
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
    ) as exc:
        raise WorkbookTemplateError("The sanitized source workbook could not be opened") from exc

    if len(workbook.worksheets) != 1:
        raise WorkbookTemplateError("The sanitized source workbook must contain one worksheet")
    if any(
        worksheet.max_row > 256 or worksheet.max_column > 256
        for worksheet in workbook.worksheets
    ):
        raise WorkbookTemplateError("The source worksheet dimensions exceed the safe limit")
    if template.worksheet_name not in workbook.sheetnames:
        raise WorkbookTemplateError("The source worksheet is missing")
    if template.worksheet_name.casefold() in AUDIT_SHEET_NAMES_CASEFOLD:
        raise WorkbookTemplateError("The source worksheet uses a reserved audit-sheet name")

    schedule = workbook[template.worksheet_name]
    nurse_ids = {nurse.id for nurse in problem.nurses}
    dates = set(_period_dates(problem))
    if set(template.nurse_rows) != nurse_ids:
        raise WorkbookTemplateError("The source workbook roster does not match the schedule")
    if set(template.date_columns) != dates:
        raise WorkbookTemplateError("The source workbook dates do not match the schedule")

    assignment_map = {
        (item.nurse_id, item.assignment_date): item.shift for item in assignments
    }
    expected_assignment_keys = {
        (nurse_id, assignment_date)
        for nurse_id in nurse_ids
        for assignment_date in dates
    }
    if set(assignment_map) != expected_assignment_keys:
        raise WorkbookTemplateError("The exported schedule is incomplete")

    for nurse_id, row in template.nurse_rows.items():
        for assignment_date, column in template.date_columns.items():
            schedule.cell(row=row, column=column).value = TEMPLATE_SHIFT_LABELS[
                assignment_map[(nurse_id, assignment_date)]
            ]

    for worksheet in list(workbook.worksheets):
        if worksheet.title.casefold() in AUDIT_SHEET_NAMES_CASEFOLD:
            workbook.remove(worksheet)
    return workbook, schedule


def build_workbook(
    problem: ScheduleProblem,
    assignments: list[Assignment],
    summaries: list[NurseSummary],
    validation: ValidationReport,
    source_workbook_template: SourceWorkbookTemplate | None = None,
) -> bytes:
    dates = _period_dates(problem)
    assignment_map = {
        (item.nurse_id, item.assignment_date): item.shift for item in assignments
    }
    if source_workbook_template is not None:
        workbook, _ = _load_source_schedule(
            problem,
            assignments,
            source_workbook_template,
        )
    else:
        workbook = Workbook()
        schedule = workbook.active
        schedule.title = "Schedule"

        schedule.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(dates))
        title = schedule.cell(row=1, column=1, value=_safe_text(problem.period_name))
        title.fill = PatternFill("solid", fgColor=NAVY)
        title.font = Font(color=WHITE, bold=True, size=16)
        title.alignment = Alignment(horizontal="left", vertical="center")
        schedule.row_dimensions[1].height = 30

        headers = ["ID", "Nickname", "Skill Level"] + [
            item.strftime("%d %a") for item in dates
        ]
        for column, value in enumerate(headers, start=1):
            _header(schedule.cell(row=2, column=column, value=value))

        shift_fills = {
            ShiftCode.DAY: PALE_YELLOW,
            ShiftCode.NIGHT: PALE_BLUE,
            ShiftCode.OFF: "EEF1F4",
            ShiftCode.VACATION: PALE_TEAL,
            ShiftCode.EDUCATION: "F3EAFB",
        }
        for row, nurse in enumerate(problem.nurses, start=3):
            identity_values = [nurse.id, nurse.nickname, nurse.skill_level.value]
            for column, value in enumerate(identity_values, start=1):
                _body(schedule.cell(row=row, column=column, value=value), centered=False)
            for offset, assignment_date in enumerate(dates, start=4):
                shift = assignment_map.get((nurse.id, assignment_date))
                cell = schedule.cell(
                    row=row,
                    column=offset,
                    value=shift.value if shift is not None else "",
                )
                _body(cell, fill=shift_fills.get(shift))
                cell.font = Font(bold=shift in {ShiftCode.DAY, ShiftCode.NIGHT})

        schedule.freeze_panes = "D3"
        schedule.auto_filter.ref = (
            f"A2:{get_column_letter(3 + len(dates))}{2 + len(problem.nurses)}"
        )
        schedule.column_dimensions["A"].width = 10
        schedule.column_dimensions["B"].width = 15
        schedule.column_dimensions["C"].width = 17
        for column in range(4, 4 + len(dates)):
            schedule.column_dimensions[get_column_letter(column)].width = 8
        schedule.sheet_view.showGridLines = False

    summary_sheet = workbook.create_sheet("Summary")
    summary_headers = [
        "ID",
        "Nickname",
        "Skill Level",
        "Day",
        "Night",
        "OFF",
        "Vacation",
        "Education",
        "Clinical Shifts",
        "Weekend Shifts",
        "Total Hours",
    ]
    for column, value in enumerate(summary_headers, start=1):
        _header(summary_sheet.cell(row=1, column=column, value=value), TEAL)
    for row, item in enumerate(summaries, start=2):
        values = [
            item.nurse_id,
            item.nickname,
            item.skill_level.value,
            item.day_count,
            item.night_count,
            item.off_count,
            item.vacation_count,
            item.education_count,
            item.clinical_shift_count,
            item.weekend_clinical_count,
            item.total_hours,
        ]
        for column, value in enumerate(values, start=1):
            _body(summary_sheet.cell(row=row, column=column, value=value))
    summary_sheet.freeze_panes = "A2"
    for column in range(1, len(summary_headers) + 1):
        summary_sheet.column_dimensions[get_column_letter(column)].width = 18
    summary_sheet.sheet_view.showGridLines = False

    validation_sheet = workbook.create_sheet("Validation")
    validation_headers = ["Constraint", "Name", "Status", "Violations", "Details"]
    for column, value in enumerate(validation_headers, start=1):
        _header(validation_sheet.cell(row=1, column=column, value=value), NAVY)
    for row, item in enumerate(validation.checks, start=2):
        values = [
            item.code,
            item.name,
            item.status,
            item.violation_count,
            "\n".join(item.details)[:32_000],
        ]
        fill = PALE_TEAL if item.status == "PASS" else PALE_RED
        for column, value in enumerate(values, start=1):
            _body(
                validation_sheet.cell(row=row, column=column, value=value),
                fill=fill,
                centered=column != 5,
            )
    validation_sheet.freeze_panes = "A2"
    validation_sheet.column_dimensions["A"].width = 34
    validation_sheet.column_dimensions["B"].width = 46
    validation_sheet.column_dimensions["C"].width = 12
    validation_sheet.column_dimensions["D"].width = 12
    validation_sheet.column_dimensions["E"].width = 90
    validation_sheet.sheet_view.showGridLines = False

    unfulfilled = workbook.create_sheet("Unfulfilled Requests")
    unfulfilled_headers = [
        "ID",
        "Nickname",
        "Date",
        "Request",
        "Priority",
        "Assigned",
    ]
    for column, value in enumerate(unfulfilled_headers, start=1):
        _header(unfulfilled.cell(row=1, column=column, value=value), TEAL)
    nurse_by_id = {nurse.id: nurse for nurse in problem.nurses}
    row = 2
    for request in sorted(problem.requests, key=lambda item: (item.request_date, item.nurse_id)):
        normalized = normalize_request(request)
        if (
            normalized.constraint_mode != RequestConstraintMode.PREFERENCE
            or not normalized.allowed_assignments
        ):
            continue
        assigned = assignment_map.get((request.nurse_id, request.request_date))
        if assigned in normalized.allowed_assignments:
            continue
        nurse = nurse_by_id[request.nurse_id]
        values = [
            nurse.id,
            nurse.nickname,
            request.request_date.isoformat(),
            request.raw_value,
            normalized.off_priority,
            assigned.value if assigned else "MISSING",
        ]
        for column, value in enumerate(values, start=1):
            _body(unfulfilled.cell(row=row, column=column, value=value))
        row += 1
    for column in range(1, len(unfulfilled_headers) + 1):
        unfulfilled.column_dimensions[get_column_letter(column)].width = 20
    unfulfilled.freeze_panes = "A2"
    unfulfilled.sheet_view.showGridLines = False

    metadata = workbook.create_sheet("Metadata")
    metadata_rows = [
        ("Period", problem.period_name),
        ("Start", problem.period_start.isoformat()),
        ("End", problem.period_end.isoformat()),
        ("Nurse count", len(problem.nurses)),
        ("Validation", "PASS" if validation.is_valid else "FAIL"),
        ("Random seed", problem.random_seed),
        ("PII policy", "Nickname and pseudonymous internal ID only"),
    ]
    for row, (label, value) in enumerate(metadata_rows, start=1):
        _header(metadata.cell(row=row, column=1, value=label), NAVY)
        _body(metadata.cell(row=row, column=2, value=value), centered=False)
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 58
    metadata.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
